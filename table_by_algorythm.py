import string
import openpyxl
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import PatternFill, Font, Alignment, numbers


def new_table_by_algorythm(path_to_params: str, headers: list[str], table: list[dict]):
    """ Function reads algorythm and makes new table """
    wb: Workbook = openpyxl.load_workbook(filename=path_to_params, data_only=True)
    for ws in wb.worksheets:
        ws: Worksheet
        if ws.title == 'Алгоритм':
            break

    # Сохраняем номера столбцов по цехам в переменные
    for col, data in enumerate(ws[2], 0):
        match data.value.lower().strip():
            case 'пекарский цех':
                baker_col: int = col
            case 'кондитерский цех':
                confectioner_col: int = col
            case 'цех слойки':
                layer_col: int = col
            case 'цех пф':
                semi_prod_col: int = col

    # Просчитываем значения по месяцам
    baker_data: dict = {}
    confectioner_data: dict = {}
    layer_data: dict = {}
    semi_prod_data: dict = {}

    for month_key in headers[1:]:
        # print(month_key)
        baker_data[month_key]: dict = {}
        confectioner_data[month_key]: dict = {}
        layer_data[month_key]: dict = {}
        semi_prod_data[month_key]: dict = {}
        for data in ws.iter_rows(min_row=3, values_only=True):
            data = list(data)
            bu_value: str = data[0].lower()

            # заполняем пустые поля в таблице "Алгоритм"
            if not data[1]:
                data[1]: float = 1.0
            for index in range(3, 7):
                if not data[index]:
                    data[index]: float = 0.0

            # Определяем соответствующую БЮ строку таблицы
            for table_row in table:
                if table_row['БЮ'].lower() == bu_value:
                    break

            baker_data[month_key][data[2]] = \
                round(baker_data[month_key].get(data[2], 0) + (data[1] * data[baker_col] * table_row[month_key]), 2)

            confectioner_data[month_key][data[2]] = \
                round(confectioner_data[month_key].get(data[2], 0) + (data[1] * data[confectioner_col] * table_row[month_key]), 2)

            layer_data[month_key][data[2]] = \
                round(layer_data[month_key].get(data[2], 0) + (data[1] * data[layer_col] * table_row[month_key]), 2)

            semi_prod_data[month_key][data[2]] = \
                round(semi_prod_data[month_key].get(data[2], 0) + (data[1] * data[semi_prod_col] * table_row[month_key]), 2)

    result: dict = {
        'пекарский цех': baker_data,
        'кондитерский цех': confectioner_data,
        'цех слойки': layer_data,
        'цех пф': semi_prod_data
    }

    return result


def save_data_to_table(source_tables: dict[dict], margin_tables: dict[dict]) -> None:
    """ Save to xslx-file """
    wb: Workbook = Workbook()
    ws: Worksheet = wb.active

    table_font: Font = Font(name='Arial', size=10)
    table_font_bold: Font = Font(name='Arial', size=10, bold=True)
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 26
    for letter in string.ascii_uppercase[2:]:
        ws.column_dimensions[letter].width = 11.5


    margin_titles: list[str] = [
        'Выпуск в ценах перемещения',
        'FC выпущенной продукции',
        'Маржа'
    ]

    sort_titles: list[str] = [
        'ФОТ+Налоги',
        'Прямые расходы',
        'Аренда помещения',
        'Электричество',
        'Питание',
        'Накладные расходы',
        'Итого расходы'
    ]

    outer_row: int = 3
    for prod_key, value in source_tables.items():
        prod_key: str
        ws.cell(row=outer_row, column=1).value = prod_key.title() if prod_key != 'цех пф' else 'Цех ПФ'
        ws.cell(row=outer_row, column=1).font = table_font_bold
        ws.cell(row=outer_row, column=1).alignment = Alignment(horizontal='center')
        col: int = 3
        for month, inner_value in value.items():
            ws.cell(row=outer_row - 1, column=col).value = month
            ws.cell(row=outer_row - 1, column=col).font = table_font_bold
            ws.cell(row=outer_row - 1, column=col).alignment = Alignment(horizontal='center')

            # Записываем строки с маржой
            margin: int = 0
            for row, title in zip(range(outer_row, outer_row + 3), margin_titles):
                if title.startswith('Выпуск'):
                    temp_num1: int|float = margin_tables[prod_key][month]['Выпуск']
                    ws.cell(row=row, column=2).value = title
                    ws.cell(row=row, column=col).value = temp_num1
                    ws.cell(row=row, column=2).font = table_font
                    ws.cell(row=row, column=col).font = table_font
                    ws.cell(row=row, column=col).number_format = '# ### ###'
                    margin += temp_num1
                elif title.startswith('FC'):
                    temp_num2: int | float = margin_tables[prod_key][month]['FC']
                    ws.cell(row=row, column=2).value = 'FC выпущенной продукции'
                    ws.cell(row=row, column=col).value = temp_num2
                    ws.cell(row=row, column=2).font = table_font
                    ws.cell(row=row, column=col).font = table_font
                    ws.cell(row=row, column=col).number_format = '# ### ###'
                    margin += temp_num2
                elif title.startswith('Маржа'):
                    ws.cell(row=row, column=2).value = 'Маржа'
                    ws.cell(row=row, column=col).value = margin
                    ws.cell(row=row, column=2).font = table_font
                    ws.cell(row=row, column=col).font = table_font
                    ws.cell(row=row, column=2).fill = PatternFill(fill_type='solid', fgColor='c2f1c8')
                    ws.cell(row=row, column=col).fill = PatternFill(fill_type='solid', fgColor='c2f1c8')
                    ws.cell(row=row, column=col).number_format = '# ### ###'

            # Записываем строки с основными расходами
            total_amount: float = 0.0
            for row, title in zip(range(outer_row + 4, outer_row + 11), sort_titles):
                if title.startswith('Итого'):
                    ws.cell(row=row, column=2).value = title
                    ws.cell(row=row, column=col).value = total_amount
                    ws.cell(row=row, column=2).font = table_font
                    ws.cell(row=row, column=col).font = table_font_bold
                    ws.cell(row=row, column=2).fill = PatternFill(fill_type='solid', fgColor='f2cfee')
                    ws.cell(row=row, column=col).fill = PatternFill(fill_type='solid', fgColor='f2cfee')
                    ws.cell(row=row, column=col).number_format = '# ### ###'
                else:
                    ws.cell(row=row, column=2).value = title
                    ws.cell(row=row, column=col).value = inner_value[title]
                    ws.cell(row=row, column=2).font = table_font
                    ws.cell(row=row, column=col).font = table_font
                    ws.cell(row=row, column=col).number_format = '# ### ###'
                    total_amount += inner_value[title]

            ws.cell(row=outer_row + 12, column=2).value = 'Прибыль цеха'
            ws.cell(row=outer_row + 12, column=col).value = total_amount + margin
            ws.cell(row=outer_row + 12, column=2).font = table_font_bold
            ws.cell(row=outer_row + 12, column=col).font = table_font_bold
            ws.cell(row=outer_row + 12, column=2).fill = PatternFill(fill_type='solid', fgColor='d9f2d0')
            ws.cell(row=outer_row + 12, column=col).fill = PatternFill(fill_type='solid', fgColor='d9f2d0')
            ws.cell(row=outer_row + 12, column=col).number_format = '# ### ###'

            col += 1
        outer_row += 15

    wb.save('results/Предварительный вариант.xlsx')


if __name__ == '__main__':
    from parameters.load_parameters import load_params
    from table_handler import set_bu_values, convert_table_to_value
    from margin_handler import margin_handler

    divisions, del_items, rules, corrections = load_params('parameters/Параметры.xlsx')

    handled_table: list[list] = set_bu_values(
        source_file='Данные для производства 01.25.xlsx',
        subdivisions=divisions,
        delete_items=del_items,
        handle_rules=rules,
        save_to_temp_file=False
    )

    table_headers, table_data = convert_table_to_value(handled_table, save_table_to_file=False)

    result_dict = new_table_by_algorythm(path_to_params='parameters/Параметры.xlsx', headers=table_headers, table=table_data)

    margin_dict = margin_handler(path_to_file='Маржа.xlsx', corrections=corrections)

    save_data_to_table(result_dict, margin_dict)