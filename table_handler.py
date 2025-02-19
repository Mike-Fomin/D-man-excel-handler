import openpyxl
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet


def set_bu_values(
        source_file: str,
        subdivisions: list[str],
        delete_items: list[str],
        handle_rules: list[dict],
        save_to_temp_file=False
) -> list[list]:
    """ Function handles source table with "БЮ" values """
    # Читаем файл-исходник
    wb: Workbook = openpyxl.load_workbook(filename=source_file, data_only=True)
    ws: Worksheet = wb.active

    # Получаем переменные: номера строки начала таблицы, столбца "Подразделение2", столбца "Статья2" и столбца "БЮ"
    # также получаем заголовки таблицы в отдельную переменную
    for header_row, data in enumerate(ws.iter_rows(values_only=True), 1):
        data: tuple[None | str | int | float]
        if data[1]:
            title_row: int = header_row
            for col, datacol in enumerate(data, 0):
                match datacol:
                    case 'Подразделение2':
                        subdivision2_col: int = col
                    case 'Статья2':
                        expence_item2_col: int = col
                    case 'БЮ':
                        bu_col: int = col
            table_headers: list[str] = list(data)
            break

    # Создание очищенной таблицы
    new_table: list = []
    for row, data in enumerate(ws.iter_rows(values_only=True, min_row=title_row + 1), title_row + 1):
        data: tuple[None | str | int | float]
        if data[subdivision2_col].lower() in subdivisions and \
                not any(map(lambda x: x in data[expence_item2_col].lower(), delete_items)):
            new_table.append(list(data))
    table_end_row: int = row

    # Обработка таблицы с учетом правил, заполнение столбца БЮ
    for rule_row in handle_rules:
        rule_row: dict[str]
        for item in new_table:
            item: list[None | str | int | float]
            if rule_row['Статья2']:
                if item[subdivision2_col].lower() == rule_row['Подразделение2'].lower() and \
                        item[expence_item2_col].lower() == rule_row['Статья2'].lower():
                    item[bu_col] = rule_row['БЮ']
            else:
                if item[subdivision2_col].lower() == rule_row['Подразделение2'].lower():
                    item[bu_col] = rule_row['БЮ']

    # Перезапись таблицы
    if save_to_temp_file:
        ws.delete_rows(idx=title_row + 1, amount=table_end_row - title_row)
        for new_row, table_row in enumerate(new_table, title_row + 1):
            for new_col, table_item in enumerate(table_row, 1):
                ws.cell(row=new_row, column=new_col).value = table_item
        wb.save('results/Таблица БЮ.xlsx')

    # Добавляем заголовки в таблицу
    new_table.insert(0, table_headers)
    # Возврат новой таблицы с удалением всех ненужных столбцов, включая столбец "сумма"
    return list(map(lambda x: [x[bu_col]] + x[bu_col + 2:], new_table))


def convert_table_to_value(table: list[list], save_table_to_file=False) -> tuple[list[str],list[dict]]:
    """ Function converts table to dict """
    table_of_dicts: list = []
    headers, table_data = table[0], [list(map(lambda x: 0 if x is None else x, row)) for row in table[1:]]

    # перебираем по ключам БЮ и суммируем ячейки по столбцам
    temp_table: list = []
    for bu_value in sorted(set(map(lambda x: x[0], table_data)), key=str.lower):
        temp_value: list = []
        for table_row in table_data:
            if table_row[0] == bu_value:
                temp_value.append(table_row)

        temp_line: list = []
        for val in zip(*temp_value):
            if isinstance(val[0], str):
                temp_line.append(bu_value)
            else:
                temp_line.append(round(sum(val), 2))
        temp_table.append(temp_line)

    # запись временной таблицы в файл
    if save_table_to_file:
        new_wb: Workbook = Workbook()
        new_ws: Worksheet = new_wb.active

        for col, header in enumerate(headers, 1):
            new_ws.cell(row=1, column=col).value = header
        for row, line in enumerate(temp_table, 2):
            for col, cell_data in enumerate(line, 1):
                if isinstance(cell_data, str):
                    new_ws.cell(row=row, column=col).value = cell_data
                else:
                    new_ws.cell(row=row, column=col).value = f"{cell_data:.2f}".replace('.', ',') if cell_data else None

        new_wb.save('results/Переходной вариант таблицы.xlsx')

    # собираем новую переменную из временной таблицы
    for table_line in temp_table:
        line_dict: dict = {key: value for key, value in zip(headers, table_line)}
        table_of_dicts.append(line_dict)

    return headers, table_of_dicts


if __name__ == '__main__':
    from parameters.load_parameters import load_params

    divisions, del_items, rules = load_params('parameters/Параметры.xlsx')

    handled_table: list[list] = set_bu_values(
        source_file='Исходник.xlsx',
        subdivisions=divisions,
        delete_items=del_items,
        handle_rules=rules,
        save_to_temp_file=False
    )

    convert_table_to_value(handled_table, save_table_to_file=False)
