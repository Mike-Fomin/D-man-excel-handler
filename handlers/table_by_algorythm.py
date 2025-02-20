import string
import openpyxl
from datetime import datetime as dt

from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import PatternFill, Font, Alignment, numbers


def new_table_by_algorythm(path_to_params: str, headers: list[str], table: list[dict]):
    """ Функция создает новую таблицу согласно алгоритму в параметрах """
    wb: Workbook = openpyxl.load_workbook(filename=path_to_params, data_only=True)
    for ws in wb.worksheets:
        ws: Worksheet
        if ws.title == 'Алгоритм':
            break

    # Сохраняем номера столбцов по цехам в переменные
    for col, data in enumerate(ws[2], 0):
        match data.value.lower().strip():
            case 'пекарский цех':
                baker_col: int = col
            case 'кондитерский цех':
                confectioner_col: int = col
            case 'цех слойки':
                layer_col: int = col
            case 'цех пф':
                semi_prod_col: int = col

    # Просчитываем значения по месяцам
    baker_data: dict = {}
    confectioner_data: dict = {}
    layer_data: dict = {}
    semi_prod_data: dict = {}
    total_guilds_data: dict = {}

    for month_key in headers[1:]:
        # print(month_key)
        baker_data[month_key]: dict = {}
        confectioner_data[month_key]: dict = {}
        layer_data[month_key]: dict = {}
        semi_prod_data[month_key]: dict = {}
        total_guilds_data[month_key]: dict = {}
        for data in ws.iter_rows(min_row=3, values_only=True):
            data = list(data)
            bu_value: str = data[0].lower()

            # заполняем пустые поля в таблице "Алгоритм"
            if not data[1]:
                data[1]: float = 1.0
            for index in range(3, 7):
                if not data[index]:
                    data[index]: float = 0.0

            # Определяем соответствующую БЮ строку таблицы
            for table_row in table:
                if table_row['БЮ'].lower() == bu_value:
                    break

            baker_data[month_key][data[2]] = \
                round(baker_data[month_key].get(data[2], 0) + (data[1] * data[baker_col] * table_row[month_key]), 2)

            confectioner_data[month_key][data[2]] = \
                round(confectioner_data[month_key].get(data[2], 0) + (data[1] * data[confectioner_col] * table_row[month_key]), 2)

            layer_data[month_key][data[2]] = \
                round(layer_data[month_key].get(data[2], 0) + (data[1] * data[layer_col] * table_row[month_key]), 2)

            semi_prod_data[month_key][data[2]] = \
                round(semi_prod_data[month_key].get(data[2], 0) + (data[1] * data[semi_prod_col] * table_row[month_key]), 2)

            total_guilds_data[month_key][data[2]] = 0

    for month in total_guilds_data:
        for unit in total_guilds_data[month]:
            total_guilds_data[month][unit] = round(sum([
                baker_data[month][unit],
                confectioner_data[month][unit],
                layer_data[month][unit],
                semi_prod_data[month][unit]
            ]), 2)

    result: dict = {
        'пекарский цех': baker_data,
        'кондитерский цех': confectioner_data,
        'цех слойки': layer_data,
        'цех пф': semi_prod_data,
        'итого': total_guilds_data
    }

    return result


def set_percents(ws: Worksheet, row: int, column: int, number: int|float, value: float| int, bold=False) -> None:
    """ Set cell value and styles"""
    table_fond_italic: Font = Font(name='Arial', size=10, italic=True, color='808080')
    table_font_italic_bold: Font = Font(name='Arial', size=10, bold=True, italic=True, color='808080')
    ws.cell(row=row, column=column).value = value / number
    if bold:
        ws.cell(row=row, column=column).font = table_font_italic_bold
    else:
        ws.cell(row=row, column=column).font = table_fond_italic
    ws.cell(row=row, column=column).number_format = '0.0%'


def save_data_to_table(
        wb: Workbook,
        source_tables: dict[dict],
        margin_tables: dict[dict],
        *,
        extended_version: bool
) -> None:
    """ Функция сохраняет новую таблицу в файл """

    if extended_version:
        ws: Worksheet = wb.create_sheet('Расширенная')
        wb.active = len(wb.worksheets) - 1
    else:
        ws: Worksheet = wb.create_sheet('Первая')
        wb.active = len(wb.worksheets) - 1

    table_font: Font = Font(name='Arial', size=10)
    table_font_bold: Font = Font(name='Arial', size=10, bold=True)

    ws.column_dimensions['A'].width = 23
    ws.column_dimensions['B'].width = 32.5
    for num, letter in enumerate(string.ascii_uppercase[2:], 2):
        if extended_version:
            if num % 2:
                ws.column_dimensions[letter].width = 9
            else:
                ws.column_dimensions[letter].width = 14
        else:
            ws.column_dimensions[letter].width = 14

    for num, letter in enumerate(string.ascii_uppercase, 0):
        if extended_version:
            if num % 2:
                ws.column_dimensions[f"A{letter}"].width = 9
            else:
                ws.column_dimensions[f"A{letter}"].width = 14
        else:
            ws.column_dimensions[f"A{letter}"].width = 14


    margin_titles: list[str] = [
        'Выпуск в ценах перемещения',
        'FC выпущенной продукции',
        'Маржа'
    ]

    sort_titles: list[str] = [
        'ФОТ+Налоги',
        'Прямые расходы',
        'Аренда помещения',
        'Электричество',
        'Питание',
        'Накладные расходы',
        'Итого расходы'
    ]

    outer_row: int = 3
    for prod_key, value in source_tables.items():
        prod_key: str
        ws.cell(row=outer_row, column=1).value = prod_key.title() if prod_key != 'цех пф' else 'Цех ПФ'
        ws.cell(row=outer_row, column=1).font = table_font_bold
        ws.cell(row=outer_row, column=1).alignment = Alignment(horizontal='center')

        col: int = 3
        if prod_key == 'итого':
            ws.cell(row=outer_row - 1, column=col - 1).fill = PatternFill(fill_type='solid', fgColor='ffff00')

        for month, inner_value in value.items():

            ws.cell(row=outer_row - 1, column=col).value = dt.strptime(month, '%d.%m.%y')
            ws.cell(row=outer_row - 1, column=col).font = table_font_bold
            ws.cell(row=outer_row - 1, column=col).alignment = Alignment(horizontal='center')
            ws.cell(row=outer_row - 1, column=col).number_format = 'MMM.YY'
            if prod_key == 'итого':
                ws.cell(row=outer_row - 1, column=col).fill = PatternFill(fill_type='solid', fgColor='ffff00')
                if extended_version:
                    ws.cell(row=outer_row - 1, column=col + 1).fill = PatternFill(fill_type='solid', fgColor='ffff00')

            # Записываем строки с маржой
            margin: int = 0
            for row, title in zip(range(outer_row, outer_row + 3), margin_titles):
                if title.startswith('Выпуск'):
                    temp_num1: int|float = margin_tables[prod_key][month]['Выпуск']
                    ws.cell(row=row, column=2).value = title
                    ws.cell(row=row, column=col).value = temp_num1
                    ws.cell(row=row, column=2).font = table_font
                    ws.cell(row=row, column=col).font = table_font
                    ws.cell(row=row, column=col).number_format = '# ### ###'
                    margin += temp_num1
                    if extended_version:
                        set_percents(ws, row, col + 1, temp_num1, temp_num1)

                elif title.startswith('FC'):
                    temp_num2: int | float = margin_tables[prod_key][month]['FC']
                    ws.cell(row=row, column=2).value = 'FC выпущенной продукции'
                    ws.cell(row=row, column=col).value = temp_num2
                    ws.cell(row=row, column=2).font = table_font
                    ws.cell(row=row, column=col).font = table_font
                    ws.cell(row=row, column=col).number_format = '# ### ###'
                    margin += temp_num2
                    if extended_version:
                        set_percents(ws, row, col + 1, temp_num1, temp_num2)
                elif title.startswith('Маржа'):
                    ws.cell(row=row, column=2).value = 'Маржа'
                    ws.cell(row=row, column=col).value = margin
                    ws.cell(row=row, column=2).font = table_font
                    ws.cell(row=row, column=col).font = table_font
                    ws.cell(row=row, column=2).fill = PatternFill(fill_type='solid', fgColor='c2f1c8')
                    ws.cell(row=row, column=col).fill = PatternFill(fill_type='solid', fgColor='c2f1c8')
                    ws.cell(row=row, column=col).number_format = '# ### ###'
                    if extended_version:
                        set_percents(ws, row, col + 1, temp_num1, margin)
                        ws.cell(row=row, column=col + 1).fill = PatternFill(fill_type='solid', fgColor='c2f1c8')

            set_percents(ws, outer_row + 3, col, temp_num1, temp_num2)

            # Записываем строки с основными расходами
            total_amount: float = 0.0
            for row, title in zip(range(outer_row + 4, outer_row + 11), sort_titles):
                if title.startswith('Итого'):
                    ws.cell(row=row, column=2).value = title
                    ws.cell(row=row, column=col).value = total_amount
                    ws.cell(row=row, column=2).font = table_font
                    ws.cell(row=row, column=col).font = table_font_bold
                    ws.cell(row=row, column=2).fill = PatternFill(fill_type='solid', fgColor='f2cfee')
                    ws.cell(row=row, column=col).fill = PatternFill(fill_type='solid', fgColor='f2cfee')
                    ws.cell(row=row, column=col).number_format = '# ### ###'
                    if extended_version:
                        set_percents(ws, row, col + 1, temp_num1, total_amount, bold=True)
                        ws.cell(row=row, column=col + 1).fill = PatternFill(fill_type='solid', fgColor='f2cfee')
                else:
                    ws.cell(row=row, column=2).value = title
                    ws.cell(row=row, column=col).value = inner_value[title]
                    ws.cell(row=row, column=2).font = table_font
                    ws.cell(row=row, column=col).font = table_font
                    ws.cell(row=row, column=col).number_format = '# ### ###'
                    total_amount += inner_value[title]
                    if extended_version:
                        set_percents(ws, row, col + 1, temp_num1, inner_value[title])

            ws.cell(row=outer_row + 12, column=2).value = 'Прибыль цеха'
            ws.cell(row=outer_row + 12, column=col).value = total_amount + margin
            ws.cell(row=outer_row + 12, column=2).font = table_font_bold
            ws.cell(row=outer_row + 12, column=col).font = table_font_bold
            ws.cell(row=outer_row + 12, column=2).fill = PatternFill(fill_type='solid', fgColor='d9f2d0')
            ws.cell(row=outer_row + 12, column=col).fill = PatternFill(fill_type='solid', fgColor='d9f2d0')
            ws.cell(row=outer_row + 12, column=col).number_format = '# ### ###'
            if extended_version:
                set_percents(ws, outer_row + 12, col + 1, temp_num1, total_amount + margin, bold=True)
                ws.cell(row=outer_row + 12, column=col + 1).fill = PatternFill(fill_type='solid', fgColor='d9f2d0')
                col += 2
            else:
                col += 1

        outer_row += 16

    if extended_version:
        wb.save('results/Результат.xlsx')
    else:
        wb.save('results/Результат.xlsx')


if __name__ == '__main__':
    pass