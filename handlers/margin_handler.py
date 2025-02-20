import openpyxl
import re

from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from parameters.load_parameters import load_params


def margin_handler(path_to_file: str, corrections: dict):
    """ Функция считывает файл Маржа.xlsx и получает из него данные"""
    wb: Workbook = openpyxl.load_workbook(filename=path_to_file, read_only=True)
    ws: Worksheet = wb.active

    # Создаем отдельный словарь для каждого цеха
    baker_data: dict = {}
    confectioner_data: dict = {}
    layer_data: dict = {}
    semi_prod_data: dict = {}
    total_guilds: dict = {}

    break_flag: bool = False
    for curr_row, row_data in enumerate(ws.iter_rows(min_row=3, values_only=True), 3):
        if break_flag:
            break
        for col, cell_data in enumerate(row_data, 0):
            if cell_data:
                if cell_data.lower() == 'цех':
                    guild_col: int = col
                elif month := re.search(pattern=r'\d{2}\.\d{2}.\d{2,4}', string=cell_data):
                    month_key: str = month.string
                    baker_data[month_key]: dict = {'FC': col + 2, 'Выпуск': col + 3}
                    confectioner_data[month_key]: dict = {'FC': col + 2, 'Выпуск': col + 3}
                    layer_data[month_key]: dict = {'FC': col + 2, 'Выпуск': col + 3}
                    semi_prod_data[month_key]: dict = {'FC': col + 2, 'Выпуск': col + 3}
                    total_guilds[month_key]: dict = {'FC': 0, 'Выпуск': 0}
                elif cell_data.lower().startswith(('горячий', 'кондитерский', 'пекарский', 'цех слойки')):
                    break_flag: bool = True
                    table_start_row: int = curr_row
                    break

    # Заполняем словарь значениями
    for row_data in ws.iter_rows(min_row=table_start_row, values_only=True):
        if row_data[guild_col]:
            match row_data[guild_col].lower().strip():
                case 'горячий цех':
                    for data_key, value in semi_prod_data.items():
                        value['FC'] = -row_data[value['FC']]
                        value['Выпуск'] = row_data[value['Выпуск']]
                case 'кондитерский цех':
                    for data_key, value in confectioner_data.items():
                        value['FC'] = -row_data[value['FC']]
                        value['Выпуск'] = row_data[value['Выпуск']]
                case 'пекарский цех':
                    for data_key, value in baker_data.items():
                        value['FC'] = -row_data[value['FC']]
                        value['Выпуск'] = row_data[value['Выпуск']]
                case 'цех слойки':
                    for data_key, value in layer_data.items():
                        value['FC'] = -row_data[value['FC']]
                        value['Выпуск'] = row_data[value['Выпуск']] - corrections.get(data_key, 0)

    for month_key in total_guilds:
        for guild in [baker_data, confectioner_data, layer_data, semi_prod_data]:
            total_guilds[month_key]['FC'] += guild[month_key]['FC']
            total_guilds[month_key]['Выпуск'] += guild[month_key]['Выпуск']

    return {
        'пекарский цех': baker_data,
        'кондитерский цех': confectioner_data,
        'цех слойки': layer_data,
        'цех пф': semi_prod_data,
        'итого': total_guilds
    }


if __name__ == '__main__':
    from pprint import pprint
    from parameters.load_parameters import load_params

    _, _ , _, corrects = load_params('../parameters/Параметры.xlsx')

    pprint(margin_handler(path_to_file='../Маржа.xlsx', corrections=corrects), sort_dicts=False)