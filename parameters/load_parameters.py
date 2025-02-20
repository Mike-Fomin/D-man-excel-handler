import openpyxl

from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet


def load_params(path_to_file: str) -> tuple[list[str], list[str], list[dict], dict]:
    """ Функция считывает параметры из файла параметров и сохраняет их в переменные """
    wb: Workbook = openpyxl.load_workbook(filename=path_to_file, data_only=True)
    for ws in wb.worksheets:
        ws: Worksheet
        match ws.title:
            case 'Подразделение2':
                subdivision: list[str] = [row_data[0].lower() for row_data in ws.iter_rows(min_row=1, values_only=True)]
            case 'Статья2':
                delete_items: list[str] = [row_data[0].lower() for row_data in ws.iter_rows(min_row=1, values_only=True)]
            case 'Правило':
                rules: list[dict] = [
                    {'БЮ': rd[0], 'Подразделение2': rd[1], 'Статья2': rd[2]}
                    for rd in ws.iter_rows(min_row=2, min_col=2, values_only=True)
                ]
            case 'Корректировка':
                corrections: dict = {key: val for key, val in zip(
                    list(map(lambda x: x.value, ws[3])),
                    list(map(lambda x: x.value, ws[4]))
                ) if key and val}

    return subdivision, delete_items, rules, corrections


if __name__ == '__main__':
    print(*load_params('Параметры.xlsx'), sep='\n')